"""
ALD_LHAR_JV
Version: 1.8
Note: this version includes PD% at 90%, 80%, 70%, 60%, 50%, 40%, 30%, 20%, 10% and user defined.
From v1.7 to v1.8: 1) profile growth vs distance, 2) Integral of theta, 3) Integral of growth 
Diffusion-reaction model for ALD on lateral high aspect ratio structures. 
Solution of the one-dimensional diffusion equation with surface reaction for partial pressure and surface coverage as function of time (Eqs. 10 and 16 of Ylilammi et al. https://doi.org/10.1063/1.5028178). It can be used in a wide range of diffusion regimes (Kn number from Kn<<1 to Kn>>1). For the surface reaction, the model uses Langmuir adsorption and includes a desorption term. For the calculations, the effective diffusion coefficient is assumed to be constant along the structure.   

DO NOT modify this script. Use/modify parameters in the file “Parameters_ALD_LHAR.xlsx”

Updated on April 21, 2026
@author: Jorge A. Velasco, by request of Prof. Riikka L. Puurunen (Catalysis Group, Aalto University). 
Funded by Genesis EU project (Chips JU, Horizon Europe).

"""
import numpy as np
from scipy.integrate import solve_ivp
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
import os

# Load the Excel workbook
wb = load_workbook("Parameters_ALD_LHAR.xlsx", data_only=True)  # data_only=True reads the value, not formula
# Select the active worksheet
ws = wb["LHAR_Parameters"]
# Read values from specific cells
Runcode = ws["B1"].value
H = ws["B3"].value   # Channel height (m)
W = ws["B4"].value     # Channel width (m)
AR = ws["B5"].value    # Aspect ratio (-)
c = ws["B6"].value    # Initial sticking probability (-)
Pd = ws["B7"].value  # Desorption probability (1/s)
q = ws["B8"].value    # Adsorption capacity (m-2) 
t_pulse = ws["B9"].value  # Pulse time (s) 
T = ws["B10"].value   # Temperature (K) 
pA0 = ws["B11"].value    # Partial pressure of reactant at z=0 (Pa) 
pI = ws["B12"].value    # Partial pressure of inert gas (Pa)
MA = ws["B13"].value    # Molar mass of reactant(kg/mol) 
dA = ws["B14"].value # Diameter of reactant (m) 
MI = ws["B15"].value  # Molar mass of inert gas (kg/mol)
dI = ws["B16"].value # Diameter of inert gas (m)

ScriptN = os.path.basename(__file__)

R = 8.31446  # Gas constant J/(mol*K) 
NA = 6.022E23 # Avogadro's number (mol-1)
kb = 1.38064852E-23; # Boltzmann constant (m2kg/s2K)

# ====== INITIAL CALCULATIONS ======
L = H*AR  #  Lenght of the channel
h = 2/(1/H + 1/W)    # Hydraulic diameter (m)
vA = np.sqrt((8*R*T)/(np.pi*MA)) # Average speed of A molecule (m/s)
DKn = (1/3)*vA*h             #
zA = np.pi/4*(dA+dI)**2*((8*R*T/np.pi*(1/MA + 1/MI))**(1/2))*pI*NA/(R*T)+np.pi*(dA)**2*((16*R*T/(np.pi*MA))**(1/2))*pA0*NA/(R*T)
DA = 3*np.pi*(vA**2)/(16*zA)  # Gas-phase diffusion constant of A molecules (m2/s)
D = 1/(1/DA + 1/DKn) # Effective diffusion constant (m2/s)
Cads = c*NA/np.sqrt(2*np.pi*MA*R*T) # Constant term for fads = (1-theta)*Cads*PA
Cdes = q*Pd    # Constant term for fdes = Cdes*theta

# Knudsen number calculations
sAA = np.pi*((dA/2)+(dA/2))**2
sAB = np.pi*((dA/2)+(dI/2))**2
term1 = (2**0.5)*sAA*pA0
term2 = ((1 + (MA/MI))**0.5)*pI*sAB
MFP = kb*T/(term1 + term2)
Knudsen = MFP/h
# Thiele modulus calculations
Alph = (L**2)*c*vA/(2*H*D);
Thiele = (Alph)**0.5;

# ====== SOLVING PDE's ======
# Space domain
Nx = ws["B17"].value            # Number of spatial grid points 
dx = L / (Nx - 1)
x = np.linspace(0, L, Nx)
# Time domain
Nt = ws["B18"].value # Number of time grid points
t_span = (0, t_pulse)
t_eval = np.linspace(*t_span, Nt)

# Initial conditions
PA_ini = np.zeros(Nx)  # Initial pressure of A along the channel
PA_ini[0] = pA0  # enforce boundary condition in initial condition
Theta_ini = np.zeros(Nx)    # Initial surface coverage along the channel
y0 = np.concatenate([PA_ini, Theta_ini])

def system(t, y):
    PA = y[:Nx]
    Theta = y[Nx:]
    
    #Theta = np.clip(Theta, 0.0, 1.0)
    
    d2PAdx2 = np.zeros(Nx)
    d2PAdx2[1:-1] = (PA[2:] - 2 * PA[1:-1] + PA[:-2]) / dx**2
    # Boundary conditions
    d2PAdx2[0] = 0       # won't be used  
    d2PAdx2[-1] = (PA[-2] - PA[-1]) / dx**2  # backward difference for Neumann BC
    dPAdt = D*d2PAdx2 - 4*R*T*(((1-Theta)*Cads*PA)-(Theta*Cdes))/(h*NA)
    dThetadt = (((1-Theta)*Cads*PA)-(Theta*Cdes))/q
    # Enforce P(x=0) = P_in for all t
    dPAdt[0] = 0
    return np.concatenate([dPAdt, dThetadt])

# Solve the coupled system
sol = solve_ivp(system, t_span, y0, t_eval=t_eval, method='BDF')  # for stiff differential equations
# Extract solutions
PA_sol = sol.y[:Nx, :]   # shape (N, len(t_eval))
Theta_sol = sol.y[Nx:, :]   # shape (N, len(t_eval))

# Extract the last time step's data
PA_final = PA_sol[:, -1]  # Last column: P(x, t_final)
Theta_final = Theta_sol[:, -1]  # Last column: T(x, t_final)

# ========== CALCULATIONS FROM FINAL PROFILE ===========

# ----- PD and slope at different theta -----

theta_add = ws["B19"].value # Additional coverage value for penetration depth and slope (-)
if theta_add > 0.99 or theta_add < 0.01:
    theta_add = 0.45

theta_targets = [0.9, 0.8, 0.7, 0.6, 0.5, 0.4, 0.3, 0.2, 0.1, theta_add] 
# Parameters values
x_at_theta = {}
x_over_H_at_theta = {}
slope_at_theta = {}
Diff_in_theta = {}

for theta_target in theta_targets:
    
    crossing_indices = np.where(np.diff(np.sign(Theta_final - theta_target)))[0]
    
    if len(crossing_indices) == 0:
        print(f"No crossing found for Theta = {theta_target}")
        
        key = round(theta_target, 2)
        x_at_theta[key] = np.nan
        x_over_H_at_theta[key] = np.nan
        slope_at_theta[key] = np.nan
        Diff_in_theta[key] = np.nan
       
        continue

    i = crossing_indices[0]
    
    Theta1, x1 = Theta_final[i], x[i]
    Theta2, x2 = Theta_final[i + 1], x[i + 1]
    
    # Interpolation
    x_interp = x1 + (theta_target - Theta1) * (x2 - x1) / (Theta2 - Theta1)
    x_over_H = x_interp / H
    
    # Slope
    slope = abs(H * (Theta2 - Theta1) / (x2 - x1))
    
    # Difference in theta (%)
    theta_diff = (Theta1 - Theta2) * 100
    
    # --- Storing data ---
    key = round(theta_target, 2)
    x_at_theta[key] = x_interp
    x_over_H_at_theta[key] = x_over_H
    slope_at_theta[key] = slope
    Diff_in_theta[key] = theta_diff

# PD and slope at theta= 0.5
PD50 = x_at_theta[0.5]/H
slopePD50 = slope_at_theta[0.5]
Backc = 13.9*slope_at_theta[0.5]**2 # for free molecular flow 
ThetaDiff = Diff_in_theta[0.5]

# Growth along the channel growth/m-2
growth = q * Theta_final

# Area under the profiles
Theta_integral = np.trapz(Theta_final, x) 
Growth_integral = np.trapz(growth, x)

# ========== PLOTS ===========

# Plot evolution of partial pressure, pA
X, Tgrid = np.meshgrid(x*1e6, sol.t) # x values in micrometers
plt.figure(figsize=(10, 5))
plt.contourf(X, Tgrid, PA_sol.T, levels=50, cmap='Blues')
plt.colorbar(label='Pressure pA(x,t)')
plt.xlabel('x (μm)')
plt.ylabel('Time (s)')
plt.title('Pressure Evolution')
plt.show()

# Plot evolution of surface coverage
plt.figure(figsize=(10, 5))
plt.contourf(X, Tgrid, Theta_sol.T, levels=50, cmap='Reds')
plt.colorbar(label='Surface coverage, theta (x,t)')
plt.xlabel('x (μm)')
plt.ylabel('Time (s)')
plt.title('Surface coverage Evolution')
plt.show()

# Plot PA(x, t_final)
plt.figure()
plt.plot(x*1e6, PA_final, label='PA(x, t_final)', color='blue') # x values in micrometers
plt.xlabel('x (μm)')
plt.ylabel('Pressure pA (Pa)')
plt.title('Final profile of partial pressure of A (pA vs x)')
plt.grid(True)
plt.legend()
plt.show()

# Plot Theta(x, t_final)
plt.figure()
plt.plot(x*1e6, Theta_final, label='Theta(x, t_final)', color='red') # x values in micrometers
plt.xlabel('x (μm)')
plt.ylabel('Surface coverage, theta (-)')
plt.title('Final surface coverage profile (theta vs x)')
plt.grid(True)
plt.legend()
plt.show()

# Plot growth (x, t_final)
plt.figure()
plt.plot(x*1e6, growth/1e18, label='growth(x, t_final)', color='green') # x values in micrometers
plt.xlabel('x (μm)')
plt.ylabel('Growth (atoms/nm^2)')
plt.title('Final growth profile (growth vs x)')
plt.grid(True)
plt.legend()
plt.show()

# Plot PA(x/H, t_final)
plt.figure()
plt.plot(x/H, PA_final, label='PA(x/H, t_final)', color='blue')
plt.xlabel('x/H (-)')
plt.ylabel('Partial pressure, pA (Pa)')
plt.title('Final profile of partial pressure of A (pA vs x/H)')
plt.grid(True)
plt.legend()
plt.show()

# Plot Theta(x/H, t_final)
plt.figure()
plt.plot(x/H, Theta_final, label='Theta(x/H, t_final)', color='red')
plt.xlabel('x/H (-)')
plt.ylabel('Surface coverage, theta (-)')
plt.title('Final surface coverage profile (theta vs x/H)')
plt.grid(True)
plt.legend()
plt.show()

# Plot growth (x/H, t_final)
plt.figure()
plt.plot(x/H, growth/1e18, label='growth(x/H, t_final)', color='green')
plt.xlabel('x/H (-)')
plt.ylabel('Growth (atoms/nm^2)')
plt.title('Final growth profile (growth vs x/H)')
plt.grid(True)
plt.legend()
plt.show()

# ====== DataFrame for exportin in Excel =====

param_data = {
    "Parameter": ["Running Code","Script's name & version:", "Channel height, H (m)", "Channel width, W (m)", "Aspect ratio, AR=L/H (-)", "Sticking coef. c (-)", "Desorption probability, Pd (1/s)", "Adsorption capacity, q (m-2)","Pulse time (s)", "T (K)", "pA0 (Pa)", "pI (Pa)", "MA (kg/mol)", "dA (m)", "MI (kg/mol)","dI (m)", "Number of spatial grid points", "Number of time grid points", "Additional θ value for reporting penetration depth and slope (-) "],
    "Value": [Runcode, ScriptN, H, W, AR, c, Pd, q, t_pulse, T, pA0, pI, MA, dA, MI, dI, Nx, Nt, theta_add]
}
param_df = pd.DataFrame(param_data)

calcvalues_data = {
   "Calculated Values": ["Lenght of Channel, L (m)", "Hydraulic diameter, h (m)","vA (m/s)", "DKn (m2/s)", "DA (m2/s)", "Deff (m2/s)", "Knudsen number (-)", "Thiele modulus (-)", "Diff. between the two data points (y-axis) around theta=0.5 (%)", "Penetration depth at 50% coverage (m)", "Penetration depth (x/H) at 50% coverage (-)", "Slope at 50% coverage, from theta vs (x/H) data, Absolute value (-)", "Back extracted 'c', only for Kn>>1 (Arts et al. 2019) (-)", "Exposure (Pa-s)", "Theta integral, ∫θ(x)dx (µm)", "Growth integral, ∫Growth(x)dx (atoms*µm/nm^2)"],
   "Value": [L, h, vA, DKn, DA, D, Knudsen, Thiele, ThetaDiff, x_at_theta[0.5], PD50, slopePD50, Backc, pA0*t_pulse, Theta_integral*1e6, Growth_integral/1e12]
}

calcvalues_df = pd.DataFrame(calcvalues_data)

data_df = pd.DataFrame({
    "x (m)": x,
    "x/H (-)": x/H,
    "pA (Pa)": PA_final,
    "theta (-)": Theta_final,
    "Growth (atoms/nm2)": growth/1e18,
})

PDvalues_df = pd.DataFrame({
    "Theta (-)": list(x_at_theta.keys()),
    "x (m)": list(x_at_theta.values()),
    "x/H (-)": list(x_over_H_at_theta.values()),
    "Slope (-)": list(slope_at_theta.values()),
    "Diff. in theta points (%)": list(Diff_in_theta.values())
})

# --- Exporting to Excel ---
run_label = Runcode
filename1 = f"{run_label}_simulation_results.xlsx"
with pd.ExcelWriter(filename1, engine="openpyxl") as writer:
    param_df.to_excel(writer, sheet_name="LHAR results", index=False, startrow=0, startcol=0)
    calcvalues_df.to_excel(writer, sheet_name="LHAR results", index=False, startrow=21, startcol=0)
    data_df.to_excel(writer, sheet_name="LHAR results", index=False, startrow=0, startcol=4)
    PDvalues_df.to_excel(writer, sheet_name="LHAR results", index=False, startrow=0, startcol=10)

    ws = writer.sheets["LHAR results"]
    for cell in ws["H"][1:]:   # skip header
        cell.number_format = '0.0000'
    for cell in ws["I"][1:]:
        cell.number_format = '0.0000'

# --- Final message ---
print(f"Excel with results saved as: {run_label}_simulation_results.xlsx")








